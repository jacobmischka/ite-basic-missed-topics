from openpyxl.utils.cell import get_column_letter

class BasicSection(object):
	HEADINGS = [
		'% total CA-1',
		'% our CA-1',
		'',
		'Difference from total',
		'Missed %'
	]

	def __init__(self, heading, subheading, items, percentages_last=False):
		self.heading = heading
		self.subheading = subheading
		self.items = [BasicItem(item, percentages_last=percentages_last) for item in items]

	def write_xlsx_headings(self, worksheet, row):
		for col, heading in enumerate(self.HEADINGS):
			worksheet.cell(row=row, column=col+3, value=heading)

	def write_xlsx_rows(self, worksheet, start_row):
		row = start_row
		worksheet.cell(row=row, column=1, value=self.heading)
		self.write_xlsx_headings(worksheet, row)

		row += 1
		subheading_row = row
		worksheet.cell(row=row, column=1, value=self.subheading)

		row += 1

		for item in self.items:
			row = item.write_xlsx_row(worksheet, row)

		BasicItem.write_xlsx_summary_row(worksheet, subheading_row, subheading_row + 1, row - 1)

		return row

	def __repr__(self):
		return 'BasicSection(heading={}, subheading={}, items={})'.format(
			self.heading,
			self.subheading,
			self.items
		)

class BasicItem(object):
	SEPARATOR_COLS = []
	KEYWORD_COL = 1

	SEPARATOR_COLS.append(2)

	TOTAL_PERCENT_COL = 3
	PERCENT_COL = 4

	SEPARATOR_COLS.append(5)

	PERCENT_DIFF_COL = 6
	MISSED_COL = 7

	def __init__(self, row, percentages_last=False):
		# Prior to 2017 the order was different on the PDF
		if percentages_last:
			(
				self.keyword,
				self.total_num,
				self.num,
				self.total_percent,
				self.percent
			) = row
		else:
			(
				self.keyword,
				self.total_num,
				self.total_percent,
				self.num,
				self.percent
			) = row

	def write_xlsx_row(self, worksheet, row):
		worksheet.cell(row=row, column=self.KEYWORD_COL, value=self.keyword)
		worksheet.cell(row=row, column=self.TOTAL_PERCENT_COL, value=self.total_percent)
		worksheet.cell(row=row, column=self.PERCENT_COL, value=self.percent)

		worksheet.cell(row=row, column=self.PERCENT_DIFF_COL,
			value='={}{} - {}{}'.format(
				get_column_letter(self.PERCENT_COL),
				row,
				get_column_letter(self.TOTAL_PERCENT_COL),
				row
			))
		worksheet.cell(row=row, column=self.MISSED_COL,
			value='=1 - {}{}'.format(
				get_column_letter(self.PERCENT_COL),
				row
			))

		return row + 1

	@classmethod
	def write_xlsx_summary_row(self, worksheet, row, start_row, end_row):
		worksheet.cell(row=row, column=self.PERCENT_DIFF_COL,
			value=self.average_column(self.PERCENT_DIFF_COL, start_row, end_row))
		worksheet.cell(row=row, column=self.MISSED_COL,
			value=self.average_column(self.MISSED_COL, start_row, end_row))

	@classmethod
	def	average_column(self, col, start_row, end_row):
		return '=AVERAGE({}{}:{}{})'.format(
			get_column_letter(col),
			start_row,
			get_column_letter(col),
			end_row
		)

	def __repr__(self):
		return (
			'BasicSection(keyword={}, total_num={}, total_percent={}, num={}, '
			'percent={})'.format(
				self.keyword,
				self.total_num,
				self.total_percent,
				self.num,
				self.percent
			)
		)
