from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.cell import get_column_letter

from .basic_section import BasicSection, BasicItem
from ..utils import get_data_ranges, get_ranges
from ..constants import (
    DIFF_GREAT_CELL,
    DIFF_GOOD_CELL,
    DIFF_BAD_CELL,
    DIFF_VERY_BAD_CELL,
    MISSED_GOOD_CELL,
    MISSED_WARNING_CELL,
    MISSED_BAD_CELL,
    DEFICIENT_DIFFERENCE_COL,
    DEFICIENT_DIFFERENCE_MISSED_COL,
    DEFICIENT_MISSED_COL,
    DIFF_GREAT,
    DIFF_GOOD,
    DIFF_BAD,
    DIFF_VERY_BAD,
    MISSED_GOOD,
    MISSED_WARNING,
    MISSED_BAD,
    DEFICIENT_DIFFERENCE,
    DEFICIENT_DIFFERENCE_MISSED,
    DEFICIENT_MISSED,
    DATA_START_ROW,
    dark_red_fill,
    light_red_fill,
    yellow_fill,
    green_fill,
    blue_fill,
)


def write_xlsx_legend(worksheet):
    worksheet[DIFF_GREAT_CELL].value = DIFF_GREAT
    worksheet[DIFF_GREAT_CELL].offset(column=-1).value = "Diff great >="

    worksheet[DIFF_GOOD_CELL].value = DIFF_GOOD
    worksheet[DIFF_GOOD_CELL].offset(column=-1).value = "Diff good >="
    worksheet[DIFF_GOOD_CELL].offset(column=2).value = "Diff warning"

    worksheet[DIFF_BAD_CELL].value = DIFF_BAD
    worksheet[DIFF_BAD_CELL].offset(column=-1).value = "Diff bad <="

    worksheet[DIFF_VERY_BAD_CELL].value = DIFF_VERY_BAD
    worksheet[DIFF_VERY_BAD_CELL].offset(column=-1).value = "Diff very bad <="

    worksheet[MISSED_GOOD_CELL].value = MISSED_GOOD
    worksheet[MISSED_GOOD_CELL].offset(column=-1).value = "Missed good <="

    worksheet[MISSED_WARNING_CELL].value = MISSED_WARNING
    worksheet[MISSED_WARNING_CELL].offset(column=-1).value = "Missed warning <="

    worksheet[MISSED_BAD_CELL].value = MISSED_BAD
    worksheet[MISSED_BAD_CELL].offset(column=-1).value = "Missed bad <="
    worksheet[MISSED_BAD_CELL].offset(column=2).value = "Missed very bad"


def dump_section_xlsx(sections, outpath):
    wb = Workbook()
    ws = wb.active

    write_xlsx_legend(ws)

    row = DATA_START_ROW
    row_ranges = []

    for section in sections:
        next_row = section.write_xlsx_rows(ws, row + 1)
        row_ranges.append((row, next_row))
        row = next_row

    data_ranges = get_data_ranges(row_ranges)
    add_styles(ws, data_ranges)
    add_conditional_formatting(ws, data_ranges)

    wb.save(outpath)


def add_styles(worksheet, data_ranges):
    subheading = NamedStyle(name="subheading")
    subheading.font = Font(name="Calibri", size=11, color="ffffff")
    subheading.fill = PatternFill(
        start_color="444444", end_color="444444", fill_type="solid"
    )
    subheading.number_format = "0%"

    for range_start, _ in data_ranges:
        for row in worksheet[
            "{}{}:{}{}".format(
                get_column_letter(BasicItem.KEYWORD_COL),
                range_start - 2,
                get_column_letter(BasicItem.MISSED_COL),
                range_start - 1,
            )
        ]:
            for cell in row:
                cell.style = subheading

    for col in BasicItem.SEPARATOR_COLS:
        for row in worksheet.iter_rows(min_row=DATA_START_ROW):
            row[col - 1].style = subheading

    column_ranges = [
        (BasicItem.TOTAL_PERCENT_COL, BasicItem.PERCENT_COL),
        (BasicItem.PERCENT_DIFF_COL, BasicItem.MISSED_COL),
    ]

    for start_col, end_col in column_ranges:
        for start_row, end_row in data_ranges:
            for row in worksheet[
                "{}{}:{}{}".format(
                    get_column_letter(start_col),
                    start_row,
                    get_column_letter(end_col),
                    end_row,
                )
            ]:
                for cell in row:
                    cell.style = "Percent"

    for cell in [
        DIFF_GREAT_CELL,
        DIFF_GOOD_CELL,
        DIFF_BAD_CELL,
        DIFF_VERY_BAD_CELL,
        MISSED_GOOD_CELL,
        MISSED_WARNING_CELL,
        MISSED_BAD_CELL,
    ]:
        worksheet[cell].style = "Percent"

    worksheet[DIFF_GREAT_CELL].offset(column=-1).fill = blue_fill
    worksheet[DIFF_GOOD_CELL].offset(column=-1).fill = green_fill
    worksheet[DIFF_GOOD_CELL].offset(column=2).fill = yellow_fill
    worksheet[DIFF_BAD_CELL].offset(column=-1).fill = light_red_fill
    worksheet[DIFF_VERY_BAD_CELL].offset(column=-1).fill = dark_red_fill

    worksheet[MISSED_GOOD_CELL].offset(column=-1).fill = green_fill
    worksheet[MISSED_WARNING_CELL].offset(column=-1).fill = yellow_fill
    worksheet[MISSED_BAD_CELL].offset(column=-1).fill = light_red_fill
    worksheet[MISSED_BAD_CELL].offset(column=2).fill = dark_red_fill


def add_conditional_formatting(worksheet, data_ranges):
    diff_range = get_ranges(BasicItem.PERCENT_DIFF_COL, data_ranges, separator=" ")
    missed_range = get_ranges(BasicItem.MISSED_COL, data_ranges, separator=" ")

    worksheet.conditional_formatting.add(
        diff_range,
        CellIsRule(operator="between", formula=[DIFF_GREAT_CELL, "1"], fill=blue_fill),
    )
    worksheet.conditional_formatting.add(
        diff_range,
        CellIsRule(
            operator="between",
            formula=[DIFF_GOOD_CELL, DIFF_GREAT_CELL],
            fill=green_fill,
        ),
    )
    worksheet.conditional_formatting.add(
        diff_range,
        CellIsRule(
            operator="between",
            formula=[DIFF_BAD_CELL, DIFF_GOOD_CELL],
            fill=yellow_fill,
        ),
    )
    worksheet.conditional_formatting.add(
        diff_range,
        CellIsRule(
            operator="between",
            formula=[DIFF_VERY_BAD_CELL, DIFF_BAD_CELL],
            fill=light_red_fill,
        ),
    )
    worksheet.conditional_formatting.add(
        diff_range,
        CellIsRule(
            operator="between", formula=["-1", DIFF_VERY_BAD_CELL], fill=dark_red_fill
        ),
    )

    worksheet.conditional_formatting.add(
        missed_range,
        CellIsRule(
            operator="between", formula=["0", MISSED_GOOD_CELL], fill=green_fill
        ),
    )
    worksheet.conditional_formatting.add(
        missed_range,
        CellIsRule(
            operator="between",
            formula=[MISSED_GOOD_CELL, MISSED_WARNING_CELL],
            fill=yellow_fill,
        ),
    )
    worksheet.conditional_formatting.add(
        missed_range,
        CellIsRule(
            operator="between",
            formula=[MISSED_WARNING_CELL, MISSED_BAD_CELL],
            fill=light_red_fill,
        ),
    )
    worksheet.conditional_formatting.add(
        missed_range,
        CellIsRule(
            operator="between", formula=[MISSED_BAD_CELL, "1"], fill=dark_red_fill
        ),
    )
