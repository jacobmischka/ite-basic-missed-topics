from openpyxl.utils.cell import get_column_letter

def get_data_ranges(row_ranges):
	return [(
		start + 3, # padding, heading, subheading
		end - 1 # padding
	) for start, end in row_ranges]

def get_range(col, start, end, end_col=None, absolute_col=False, absolute_row=False):
	if not end_col:
		end_col = col

	return '{}{}{}{}:{}{}{}{}'.format(
		'$' if absolute_col else '',
		get_column_letter(col),
		'$' if absolute_row else '',
		start,
		'$' if absolute_col else '',
		get_column_letter(end_col),
		'$' if absolute_row else '',
		end
	)

def get_range_list(col, ranges, end_col=None, absolute_col=False, absolute_row=False):
	return [
		get_range(col, start, end, end_col=end_col, absolute_col=absolute_col, absolute_row=absolute_row)
		for start, end in ranges
	]

def get_ranges(col, ranges, end_col=None, absolute_col=False, absolute_row=False, separator=','):
	return separator.join(get_range_list(col, ranges, end_col, absolute_col, absolute_row))
