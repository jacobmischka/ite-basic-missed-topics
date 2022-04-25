from openpyxl.styles import PatternFill

DIFF_GREAT_CELL = "$D$1"
DIFF_GOOD_CELL = "$G$1"
DIFF_BAD_CELL = "$L$1"
DIFF_VERY_BAD_CELL = "$O$1"

MISSED_GOOD_CELL = "$D$2"
MISSED_WARNING_CELL = "$G$2"
MISSED_BAD_CELL = "$J$2"

DEFICIENT_DIFFERENCE_COL = "$T$1"
DEFICIENT_DIFFERENCE_MISSED_COL = "$Y$1"
DEFICIENT_MISSED_COL = "$Y$2"

DIFF_GREAT = 0.2
DIFF_GOOD = 0
DIFF_BAD = -0.1
DIFF_VERY_BAD = -0.2

MISSED_GOOD = 0.25
MISSED_WARNING = 0.5
MISSED_BAD = 0.75

DEFICIENT_DIFFERENCE = -0.15
DEFICIENT_DIFFERENCE_MISSED = 0
DEFICIENT_MISSED = 1

DATA_START_ROW = 5

dark_red_fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
light_red_fill = PatternFill(
    start_color="ff7d7d", end_color="ff7d7d", fill_type="solid"
)
yellow_fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
green_fill = PatternFill(start_color="55ff55", end_color="55ff55", fill_type="solid")
blue_fill = PatternFill(start_color="00c2ff", end_color="00c2ff", fill_type="solid")
