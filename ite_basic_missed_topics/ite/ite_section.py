from openpyxl.utils.cell import get_column_letter

from ..constants import (
	DIFFICIENT_DIFFERENCE_COL,
	DIFFICIENT_DIFFERENCE_MISSED_COL,
	DIFFICIENT_MISSED_COL
)

class IteSection(object):
	HEADINGS = [
		'',
		'% total CBY',
		'% our CBY',
		'% total CA-1',
		'% our CA-1',
		'% total CA-2',
		'% our CA-2',
		'% total CA-3',
		'% our CA-3',
		'',
		'CBY',
		'CA-1',
		'CA-2',
		'CA-3',
		'',
		'CBY',
		'CA-1',
		'CA-2',
		'CA-3',
		'',
		'Advanced or Basic',
		'CBY',
		'CA-1',
		'CA-2',
		'CA-3'
	]

	def __init__(self, heading, subheading, items):
		self.heading = heading
		self.subheading = subheading
		self.items = [IteItem(item) for item in items]


	def get_csv_rows(self):
		return [
			[
				self.heading,
				*self.HEADINGS
			],
			*[item.get_csv_row() for item in self.items]
		]

	def write_xlsx_rows(self, worksheet, start_row):
		row = start_row
		worksheet.cell(row=row, column=1, value=self.heading)
		self.write_xlsx_headings(worksheet, row)

		row += 1
		subheading_row = row
		worksheet.cell(row=subheading_row, column=1, value=self.subheading)

		row += 1

		for item in self.items:
			row = item.write_xlsx_row(worksheet, row)

		IteItem.write_xlsx_summary_row(worksheet, subheading_row, subheading_row + 1, row - 1)

		return row

	@classmethod
	def write_xlsx_headings(self, worksheet, row):
		for col, heading in enumerate(self.HEADINGS):
			worksheet.cell(row=row, column=col+2, value=heading)

	def __repr__(self):
		return 'IteSection(heading={}, subheading={}, items={})'.format(
			self.heading,
			self.subheading,
			self.items
		)

class IteItem(object):
	SEPARATOR_COLS = []

	KEYWORD_COL = 1

	SEPARATOR_COLS.append(2)

	CBY_TOTAL_COL = 3
	CBY_COL = 4
	CA1_TOTAL_COL = 5
	CA1_COL = 6
	CA2_TOTAL_COL = 7
	CA2_COL = 8
	CA3_TOTAL_COL = 9
	CA3_COL = 10

	SEPARATOR_COLS.append(11)

	CBY_DIFF_COL = 12
	CA1_DIFF_COL = 13
	CA2_DIFF_COL = 14
	CA3_DIFF_COL = 15

	SEPARATOR_COLS.append(16)

	CBY_MISSED_COL = 17
	CA1_MISSED_COL = 18
	CA2_MISSED_COL = 19
	CA3_MISSED_COL = 20

	SEPARATOR_COLS.append(21)

	A_OR_B_COL = 22
	CBY_DIFFICIENT_COL = 23
	CA1_DIFFICIENT_COL = 24
	CA2_DIFFICIENT_COL = 25
	CA3_DIFFICIENT_COL = 26

	def __init__(self, row):
		self.keyword = row[0]
		(
			self.ca3_total,
			self.ca3,
			self.ca2_total,
			self.ca2,
			self.ca1_total,
			self.ca1,
			self.cby_total,
			self.cby
		) = [parse_percentage(cell) for cell in row[1:]]

		self.item_type = 'A' if '(A)' in self.keyword else 'B'

	@property
	def cby_diff(self):
		return self.cby - self.cby_total

	@property
	def ca1_diff(self):
		return self.ca1 - self.ca1_total

	@property
	def ca2_diff(self):
		return self.ca2 - self.ca2_total

	@property
	def ca3_diff(self):
		return self.ca3 - self.ca3_total

	@property
	def cby_missed(self):
		return 1 - self.cby

	@property
	def ca1_missed(self):
		return 1 - self.ca1

	@property
	def ca2_missed(self):
		return 1 - self.ca2

	@property
	def ca3_missed(self):
		return 1 - self.ca3

	def get_csv_row(self):
		return [
			self.keyword,
			'',
			self.cby_total,
			self.cby,
			self.ca1_total,
			self.ca1,
			self.ca2_total,
			self.ca2,
			self.ca3_total,
			self.ca3,
			'',
			self.cby_diff,
			self.ca1_diff,
			self.ca2_diff,
			self.ca3_diff,
			'',
			self.cby_missed,
			self.ca1_missed,
			self.ca2_missed,
			self.ca3_missed
		]

	@classmethod
	def write_xlsx_summary_row(self, worksheet, row, start_row, end_row):
		worksheet.cell(row=row, column=self.CBY_DIFF_COL,
			value=self.average_column(self.CBY_DIFF_COL, start_row, end_row))
		worksheet.cell(row=row, column=self.CA1_DIFF_COL,
			value=self.average_column(self.CA1_DIFF_COL, start_row, end_row))
		worksheet.cell(row=row, column=self.CA2_DIFF_COL,
			value=self.average_column(self.CA2_DIFF_COL, start_row, end_row))
		worksheet.cell(row=row, column=self.CA3_DIFF_COL,
			value=self.average_column(self.CA3_DIFF_COL, start_row, end_row))

		worksheet.cell(row=row, column=self.CBY_MISSED_COL,
			value=self.average_column(self.CBY_MISSED_COL, start_row, end_row))
		worksheet.cell(row=row, column=self.CA1_MISSED_COL,
			value=self.average_column(self.CA1_MISSED_COL, start_row, end_row))
		worksheet.cell(row=row, column=self.CA2_MISSED_COL,
			value=self.average_column(self.CA2_MISSED_COL, start_row, end_row))
		worksheet.cell(row=row, column=self.CA3_MISSED_COL,
			value=self.average_column(self.CA3_MISSED_COL, start_row, end_row))


	@classmethod
	def	average_column(self, col, start_row, end_row):
		return '=AVERAGE({}{}:{}{})'.format(
			get_column_letter(col),
			start_row,
			get_column_letter(col),
			end_row
		)

	def write_xlsx_row(self, worksheet, row):

		worksheet.cell(row=row, column=self.KEYWORD_COL, value=self.keyword)

		worksheet.cell(row=row, column=self.CBY_TOTAL_COL, value=self.cby_total)
		worksheet.cell(row=row, column=self.CBY_COL, value=self.cby)
		worksheet.cell(row=row, column=self.CA1_TOTAL_COL, value=self.ca1_total)
		worksheet.cell(row=row, column=self.CA1_COL, value=self.ca1)
		worksheet.cell(row=row, column=self.CA2_TOTAL_COL, value=self.ca2_total)
		worksheet.cell(row=row, column=self.CA2_COL, value=self.ca2)
		worksheet.cell(row=row, column=self.CA3_TOTAL_COL, value=self.ca3_total)
		worksheet.cell(row=row, column=self.CA3_COL, value=self.ca3)

		worksheet.cell(row=row, column=self.CBY_DIFF_COL,
			value='={}{} - {}{}'.format(
				get_column_letter(self.CBY_COL),
				row,
				get_column_letter(self.CBY_TOTAL_COL),
				row
			))
		worksheet.cell(row=row, column=self.CA1_DIFF_COL,
			value='={}{} - {}{}'.format(
				get_column_letter(self.CA1_COL),
				row,
				get_column_letter(self.CA1_TOTAL_COL),
				row
			))
		worksheet.cell(row=row, column=self.CA2_DIFF_COL,
			value='={}{} - {}{}'.format(
				get_column_letter(self.CA2_COL),
				row,
				get_column_letter(self.CA2_TOTAL_COL),
				row
			))
		worksheet.cell(row=row, column=self.CA3_DIFF_COL,
			value='={}{} - {}{}'.format(
				get_column_letter(self.CA3_COL),
				row,
				get_column_letter(self.CA3_TOTAL_COL),
				row
			))

		worksheet.cell(row=row, column=self.CBY_MISSED_COL,
			value='=1 - {}{}'.format(
				get_column_letter(self.CBY_COL),
				row
			))
		worksheet.cell(row=row, column=self.CA1_MISSED_COL,
			value='=1 - {}{}'.format(
				get_column_letter(self.CA1_COL),
				row
			))
		worksheet.cell(row=row, column=self.CA2_MISSED_COL,
			value='=1 - {}{}'.format(
				get_column_letter(self.CA2_COL),
				row
			))
		worksheet.cell(row=row, column=self.CA3_MISSED_COL,
			value='=1 - {}{}'.format(
				get_column_letter(self.CA3_COL),
				row
			))

		worksheet.cell(row=row, column=self.A_OR_B_COL, value=self.item_type)

		worksheet.cell(row=row, column=self.CBY_DIFFICIENT_COL,
			value='=IF(OR(AND({}{} >= {}, {}{} >= {}), {}{} >= {}), "YES", "")'.format(
				get_column_letter(self.CBY_DIFF_COL),
				row,
				DIFFICIENT_DIFFERENCE_COL,
				get_column_letter(self.CBY_MISSED_COL),
				row,
				DIFFICIENT_DIFFERENCE_MISSED_COL,
				get_column_letter(self.CBY_MISSED_COL),
				row,
				DIFFICIENT_MISSED_COL
			))
		worksheet.cell(row=row, column=self.CA1_DIFFICIENT_COL,
			value='=IF(OR(AND({}{} >= {}, {}{} >= {}), {}{} >= {}), "YES", "")'.format(
				get_column_letter(self.CA1_DIFF_COL),
				row,
				DIFFICIENT_DIFFERENCE_COL,
				get_column_letter(self.CA1_MISSED_COL),
				row,
				DIFFICIENT_DIFFERENCE_MISSED_COL,
				get_column_letter(self.CA1_MISSED_COL),
				row,
				DIFFICIENT_MISSED_COL
			))
		worksheet.cell(row=row, column=self.CA2_DIFFICIENT_COL,
			value='=IF(OR(AND({}{} >= {}, {}{} >= {}), {}{} >= {}), "YES", "")'.format(
				get_column_letter(self.CA2_DIFF_COL),
				row,
				DIFFICIENT_DIFFERENCE_COL,
				get_column_letter(self.CA2_MISSED_COL),
				row,
				DIFFICIENT_DIFFERENCE_MISSED_COL,
				get_column_letter(self.CA2_MISSED_COL),
				row,
				DIFFICIENT_MISSED_COL
			))
		worksheet.cell(row=row, column=self.CA3_DIFFICIENT_COL,
			value='=IF(OR(AND({}{} >= {}, {}{} >= {}), {}{} >= {}), "YES", "")'.format(
				get_column_letter(self.CA3_DIFF_COL),
				row,
				DIFFICIENT_DIFFERENCE_COL,
				get_column_letter(self.CA3_MISSED_COL),
				row,
				DIFFICIENT_DIFFERENCE_MISSED_COL,
				get_column_letter(self.CA3_MISSED_COL),
				row,
				DIFFICIENT_MISSED_COL
			))


		return row + 1

	def __repr__(self):
		return (
			'IteItem(keyword={}, ca3_total={}, ca3={}, ca2_total={}, ca2={}, '
			'ca1_total={}, ca1={}, cby_total={}, cby={})'
		).format(
			self.keyword,
			self.ca3_total,
			self.ca3,
			self.ca2_total,
			self.ca2,
			self.ca1_total,
			self.ca1,
			self.cby_total,
			self.cby
		)


def parse_percentage(text):
	return float(text[:-1]) / 100

def format_percentage(num):
	return '{}%'.format(num * 100)
