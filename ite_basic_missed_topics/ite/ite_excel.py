from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.cell import get_column_letter

from .ite_section import IteSection, IteItem
from ..utils import get_data_ranges, get_range, get_ranges
from ..constants import (
	DIFF_GREAT_CELL,
	DIFF_GOOD_CELL,
	DIFF_BAD_CELL,
	DIFF_VERY_BAD_CELL,

	MISSED_GOOD_CELL,
	MISSED_WARNING_CELL,
	MISSED_BAD_CELL,

	DIFFICIENT_DIFFERENCE_COL,
	DIFFICIENT_DIFFERENCE_MISSED_COL,
	DIFFICIENT_MISSED_COL,

	DIFF_GREAT,
	DIFF_GOOD,
	DIFF_BAD,
	DIFF_VERY_BAD,

	MISSED_GOOD,
	MISSED_WARNING,
	MISSED_BAD,

	DIFFICIENT_DIFFERENCE,
	DIFFICIENT_DIFFERENCE_MISSED,
	DIFFICIENT_MISSED,

	DATA_START_ROW,

	dark_red_fill,
	light_red_fill,
	yellow_fill,
	green_fill,
	blue_fill
)

def write_xlsx_legend(worksheet):
	worksheet[DIFF_GREAT_CELL].value = DIFF_GREAT
	worksheet[DIFF_GREAT_CELL].offset(column=-1).value = 'Diff great >='

	worksheet[DIFF_GOOD_CELL].value = DIFF_GOOD
	worksheet[DIFF_GOOD_CELL].offset(column=-1).value = 'Diff good >='
	worksheet[DIFF_GOOD_CELL].offset(column=2).value = 'Diff warning'

	worksheet[DIFF_BAD_CELL].value = DIFF_BAD
	worksheet[DIFF_BAD_CELL].offset(column=-1).value = 'Diff bad <='

	worksheet[DIFF_VERY_BAD_CELL].value = DIFF_VERY_BAD
	worksheet[DIFF_VERY_BAD_CELL].offset(column=-1).value = 'Diff very bad <='

	worksheet[MISSED_GOOD_CELL].value = MISSED_GOOD
	worksheet[MISSED_GOOD_CELL].offset(column=-1).value = 'Missed good <='

	worksheet[MISSED_WARNING_CELL].value = MISSED_WARNING
	worksheet[MISSED_WARNING_CELL].offset(column=-1).value = 'Missed warning <='

	worksheet[MISSED_BAD_CELL].value = MISSED_BAD
	worksheet[MISSED_BAD_CELL].offset(column=-1).value = 'Missed bad <='
	worksheet[MISSED_BAD_CELL].offset(column=2).value = 'Missed very bad'

	worksheet[DIFFICIENT_DIFFERENCE_COL].offset(column=-2).value = 'Difficient if'
	worksheet[DIFFICIENT_DIFFERENCE_COL].offset(column=-1).value = 'National mean difference greater than'
	worksheet[DIFFICIENT_DIFFERENCE_COL].value = DIFFICIENT_DIFFERENCE

	worksheet[DIFFICIENT_DIFFERENCE_MISSED_COL].offset(column=-3).value = 'AND'
	worksheet[DIFFICIENT_DIFFERENCE_MISSED_COL].offset(column=-1).value = '% Missed greater than'
	worksheet[DIFFICIENT_DIFFERENCE_MISSED_COL].value = DIFFICIENT_DIFFERENCE_MISSED

	worksheet[DIFFICIENT_MISSED_COL].offset(column=-3).value = 'OR'
	worksheet[DIFFICIENT_MISSED_COL].offset(column=-1).value = '% Missed greater than'
	worksheet[DIFFICIENT_MISSED_COL].value = DIFFICIENT_MISSED


def dump_section_xlsx(sections, outpath):
	wb = Workbook()
	ws = wb.active

	write_xlsx_legend(ws)

	row = DATA_START_ROW
	row_ranges = []

	ws.cell(row=row, column=IteItem.CBY_TOTAL_COL, value='Original Data')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_TOTAL_COL, end_column=IteItem.CA3_COL)
	ws.cell(row=row, column=IteItem.CBY_DIFF_COL, value='Difference from National Mean')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_DIFF_COL, end_column=IteItem.CA3_DIFF_COL)
	ws.cell(row=row, column=IteItem.CBY_MISSED_COL, value='% Missed')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_MISSED_COL, end_column=IteItem.CA3_MISSED_COL)
	ws.cell(row=row, column=IteItem.CBY_DIFFICIENT_COL, value='Difficient Area')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_DIFFICIENT_COL, end_column=IteItem.CA3_DIFFICIENT_COL)

	for section in sections:
		end_row = section.write_xlsx_rows(ws, row + 1)
		row_ranges.append((row, end_row))
		row = end_row

	data_ranges = get_data_ranges(row_ranges)
	IteSection.write_xlsx_headings(ws, row + 1)
	write_xlsx_summary(ws, data_ranges, row + 2)
	add_conditional_formatting(ws, data_ranges)
	add_styles(ws, data_ranges)

	wb.save(outpath)

def add_styles(worksheet, data_ranges):
	subheading = NamedStyle(name='subheading')
	subheading.font = Font(name='Calibri', size=11, color='ffffff')
	subheading.fill = PatternFill(start_color='444444', end_color='444444',
		fill_type='solid')
	subheading.number_format = '0%'

	for range_start, _ in data_ranges:
		for row in worksheet['{}{}:{}{}'.format(
			get_column_letter(IteItem.KEYWORD_COL),
			range_start - 2,
			get_column_letter(IteItem.CA3_DIFFICIENT_COL),
			range_start - 1
		)]:
			for cell in row:
				cell.style = subheading

	# summary subheadings
	for row in worksheet['{}{}:{}{}'.format(
		get_column_letter(IteItem.KEYWORD_COL),
		data_ranges[-1][1] + 1,
		get_column_letter(IteItem.CA3_DIFFICIENT_COL),
		data_ranges[-1][1] + 2
	)]:
		for cell in row:
			cell.style = subheading

	# vertical separators
	for col in IteItem.SEPARATOR_COLS + [IteItem.A_OR_B_COL]:
		for row in worksheet.iter_rows(min_row=DATA_START_ROW):
			row[col - 1].style = subheading

	heading = NamedStyle(name='heading')
	heading.font = Font(name='Calibri', size=11, bold=True)
	heading.alignment = Alignment(horizontal='center')

	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_TOTAL_COL).style = heading
	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_DIFF_COL).style = heading
	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_MISSED_COL).style = heading
	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_DIFFICIENT_COL).style = heading

	column_ranges = [
		(IteItem.CBY_TOTAL_COL, IteItem.CA3_COL),
		(IteItem.CBY_DIFF_COL, IteItem.CA3_DIFF_COL),
		(IteItem.CBY_MISSED_COL, IteItem.CA3_MISSED_COL)
	]

	data_ranges.append([
		(data_ranges[-1][1] + 3),
		(data_ranges[-1][1] + 5)
	])

	for start_col, end_col in column_ranges:
		for start_row, end_row in data_ranges:
			for row in worksheet['{}{}:{}{}'.format(
				get_column_letter(start_col),
				start_row,
				get_column_letter(end_col),
				end_row
			)]:
				for cell in row:
					cell.style = 'Percent'

	for cell in [
		DIFF_GREAT_CELL,
		DIFF_GOOD_CELL,
		DIFF_BAD_CELL,
		DIFF_VERY_BAD_CELL,
		MISSED_GOOD_CELL,
		MISSED_WARNING_CELL,
		MISSED_BAD_CELL,
		DIFFICIENT_DIFFERENCE_COL,
		DIFFICIENT_DIFFERENCE_MISSED_COL,
		DIFFICIENT_MISSED_COL
	]:
		worksheet[cell].style = 'Percent'

	worksheet[DIFF_GREAT_CELL].offset(column=-1).fill = blue_fill
	worksheet[DIFF_GOOD_CELL].offset(column=-1).fill = green_fill
	worksheet[DIFF_GOOD_CELL].offset(column=2).fill = yellow_fill
	worksheet[DIFF_BAD_CELL].offset(column=-1).fill = light_red_fill
	worksheet[DIFF_VERY_BAD_CELL].offset(column=-1).fill = dark_red_fill

	worksheet[MISSED_GOOD_CELL].offset(column=-1).fill = green_fill
	worksheet[MISSED_WARNING_CELL].offset(column=-1).fill = yellow_fill
	worksheet[MISSED_BAD_CELL].offset(column=-1).fill = light_red_fill
	worksheet[MISSED_BAD_CELL].offset(column=2).fill = dark_red_fill

def add_conditional_formatting(worksheet, data_ranges):
	diff_range = get_ranges(IteItem.CBY_DIFF_COL, data_ranges,
		end_col=IteItem.CA3_DIFF_COL, separator=' ')
	missed_range = get_ranges(IteItem.CBY_MISSED_COL, data_ranges,
		end_col=IteItem.CA3_MISSED_COL, separator=' ')

	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_GREAT_CELL, '1'], fill=blue_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_GOOD_CELL, DIFF_GREAT_CELL], fill=green_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_BAD_CELL, DIFF_GOOD_CELL], fill=yellow_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_VERY_BAD_CELL, DIFF_BAD_CELL], fill=light_red_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=['-1', DIFF_VERY_BAD_CELL], fill=dark_red_fill))


	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=['0', MISSED_GOOD_CELL], fill=green_fill))
	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=[MISSED_GOOD_CELL, MISSED_WARNING_CELL], fill=yellow_fill))
	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=[MISSED_WARNING_CELL, MISSED_BAD_CELL], fill=light_red_fill))
	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=[MISSED_BAD_CELL, '1'], fill=dark_red_fill))

def write_xlsx_summary(worksheet, data_ranges, row):
	OVERALL_ROW = row
	ADVANCED_ROW = row + 1
	BASIC_ROW = row + 2

	worksheet.cell(row=OVERALL_ROW, column=IteItem.KEYWORD_COL, value='Overall average')
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_COL, data_ranges)))

	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_DIFF_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_DIFF_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_DIFF_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_DIFF_COL, data_ranges)))

	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_MISSED_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_MISSED_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_MISSED_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_MISSED_COL, data_ranges)))

	worksheet.cell(row=ADVANCED_ROW, column=1, value='Advanced question average')
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=BASIC_ROW, column=1, value='Advanced question average')
	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
