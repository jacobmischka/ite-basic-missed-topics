#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.cell import get_column_letter

import csv, sys

from ite_section import IteSection, IteItem
from utils import get_data_ranges, get_range, get_range_list, get_ranges

DIFF_GREAT_CELL = 'D1'
DIFF_GOOD_CELL = 'G1'
DIFF_BAD_CELL = 'J1'
DIFF_VERY_BAD_CELL = 'M1'

MISSED_GOOD_CELL = 'D2'
MISSED_WARNING_CELL = 'G2'
MISSED_BAD_CELL = 'J2'

DIFF_GREAT = 0.2
DIFF_GOOD = 0
DIFF_BAD = -0.1
DIFF_VERY_BAD = -0.2

MISSED_GOOD = 0.25
MISSED_WARNING = 0.5
MISSED_BAD = 0.75

DATA_START_ROW = 5

def extract(inpath):
	rows = []
	row = None

	with open(inpath, 'r') as infile:
		for line in infile:
			if should_skip(line):
				continue

			if is_new_row(line):
				if row:
					rows.append(row)
				row = []

			if row != None:
				row.append(line.strip())

	labels = rows[0]

	body = [row for row in rows if row != labels]

	return labels, body

def is_heading(line):
	# FIXME: This isn't good
	return (
		'Keyword' in line
		or 'Basic Sciences' in line
		or 'Clinical Sciences' in line
		or 'Clinical Subspecialties' in line
		or 'Special Problems' in line
		or 'Basic items' in line
		or 'Advanced items' in line
	)

def is_new_row(line):
	return (
		'(A)' in line
		or '(B)' in line
		or is_heading(line)
	)

def should_skip(line):
	return (
		not line
		or len(line.strip()) == 0
		or 'Page' in line
		or '#' in line
		or 'N=' in line
	)

def dump_csv(labels, body, outpath):
	rows = [
		labels,
		*body
	]

	with open(outpath, 'w') as outfile:
		writer = csv.writer(outfile)
		writer.writerows(rows)

def extract_sections(rows):
	sections = []
	heading = None
	subheading = None
	items = []

	for row in rows:
		if len(row) == 1:
			if not heading:
				heading = row[0]
			elif not subheading:
				subheading = row[0]
			elif items:
				try:
					sections.append(IteSection(heading, subheading, items))
					heading = row[0]
					subheading = None
					items = []
				except Exception as e:
					print(e, file=sys.stderr)
		else:
			items.append(row)

	try:
		sections.append(IteSection(heading, subheading, items))
	except Exception as e:
		print(e, file=sys.stderr)

	return sections

def get_csv_rows(sections):
	rows = []
	items = []

	for section in sections:
		rows += section.get_csv_rows()
		rows.append([])
		items += section.items

	rows.append([])

	rows.append([
		'Overall averages',
		'',
		sum([item.cby_total for item in items])/len(items),
		sum([item.cby for item in items])/len(items),
		sum([item.ca1_total for item in items])/len(items),
		sum([item.ca1 for item in items])/len(items),
		sum([item.ca2_total for item in items])/len(items),
		sum([item.ca2 for item in items])/len(items),
		sum([item.ca3_total for item in items])/len(items),
		sum([item.ca3 for item in items])/len(items),
		'',
		sum([item.cby_diff for item in items])/len(items),
		sum([item.ca1_diff for item in items])/len(items),
		sum([item.ca2_diff for item in items])/len(items),
		sum([item.ca3_diff for item in items])/len(items),
		'',
		sum([item.cby_missed for item in items])/len(items),
		sum([item.ca1_missed for item in items])/len(items),
		sum([item.ca2_missed for item in items])/len(items),
		sum([item.ca3_missed for item in items])/len(items)
	])

	advanced_items = [item for item in items if item.item_type == 'A']
	basic_items = [item for item in items if item.item_type == 'B']

	rows.append([
		'Advanced averages',
		'',
		sum([item.cby_total for item in advanced_items])/len(advanced_items),
		sum([item.cby for item in advanced_items])/len(advanced_items),
		sum([item.ca1_total for item in advanced_items])/len(advanced_items),
		sum([item.ca1 for item in advanced_items])/len(advanced_items),
		sum([item.ca2_total for item in advanced_items])/len(advanced_items),
		sum([item.ca2 for item in advanced_items])/len(advanced_items),
		sum([item.ca3_total for item in advanced_items])/len(advanced_items),
		sum([item.ca3 for item in advanced_items])/len(advanced_items),
		'',
		sum([item.cby_diff for item in advanced_items])/len(advanced_items),
		sum([item.ca1_diff for item in advanced_items])/len(advanced_items),
		sum([item.ca2_diff for item in advanced_items])/len(advanced_items),
		sum([item.ca3_diff for item in advanced_items])/len(advanced_items),
		'',
		sum([item.cby_missed for item in advanced_items])/len(advanced_items),
		sum([item.ca1_missed for item in advanced_items])/len(advanced_items),
		sum([item.ca2_missed for item in advanced_items])/len(advanced_items),
		sum([item.ca3_missed for item in advanced_items])/len(advanced_items)
	])
	rows.append([
		'Basic averages',
		'',
		sum([item.cby_total for item in basic_items])/len(basic_items),
		sum([item.cby for item in basic_items])/len(basic_items),
		sum([item.ca1_total for item in basic_items])/len(basic_items),
		sum([item.ca1 for item in basic_items])/len(basic_items),
		sum([item.ca2_total for item in basic_items])/len(basic_items),
		sum([item.ca2 for item in basic_items])/len(basic_items),
		sum([item.ca3_total for item in basic_items])/len(basic_items),
		sum([item.ca3 for item in basic_items])/len(basic_items),
		'',
		sum([item.cby_diff for item in basic_items])/len(basic_items),
		sum([item.ca1_diff for item in basic_items])/len(basic_items),
		sum([item.ca2_diff for item in basic_items])/len(basic_items),
		sum([item.ca3_diff for item in basic_items])/len(basic_items),
		'',
		sum([item.cby_missed for item in basic_items])/len(basic_items),
		sum([item.ca1_missed for item in basic_items])/len(basic_items),
		sum([item.ca2_missed for item in basic_items])/len(basic_items),
		sum([item.ca3_missed for item in basic_items])/len(basic_items)
	])

	return rows

def dump_section_csv(sections, outpath):
	with open(outpath, 'w') as outfile:
		writer = csv.writer(outfile)
		writer.writerows(get_csv_rows(sections))

def write_xlsx_legend(worksheet):
	worksheet[DIFF_GREAT_CELL].value = DIFF_GREAT
	worksheet[DIFF_GREAT_CELL].offset(column=-1).value = 'Diff great'

	worksheet[DIFF_GOOD_CELL].value = DIFF_GOOD
	worksheet[DIFF_GOOD_CELL].offset(column=-1).value = 'Diff good'

	worksheet[DIFF_BAD_CELL].value = DIFF_BAD
	worksheet[DIFF_BAD_CELL].offset(column=-1).value = 'Diff bad'

	worksheet[DIFF_VERY_BAD_CELL].value = DIFF_VERY_BAD
	worksheet[DIFF_VERY_BAD_CELL].offset(column=-1).value = 'Diff very bad'

	worksheet[MISSED_GOOD_CELL].value = MISSED_GOOD
	worksheet[MISSED_GOOD_CELL].offset(column=-1).value = 'Missed good'

	worksheet[MISSED_WARNING_CELL].value = MISSED_WARNING
	worksheet[MISSED_WARNING_CELL].offset(column=-1).value = 'Missed warning'

	worksheet[MISSED_BAD_CELL].value = MISSED_BAD
	worksheet[MISSED_BAD_CELL].offset(column=-1).value = 'Missed bad'


def dump_section_xlsx(sections, outpath):
	wb = Workbook()
	ws = wb.active

	write_xlsx_legend(ws)

	row = DATA_START_ROW
	row_ranges = []

	ws.cell(row=row, column=IteItem.CBY_TOTAL_COL, value='Original Data')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_TOTAL_COL, end_column=IteItem.CA3_COL)
	ws.cell(row=row, column=IteItem.CBY_DIFF_COL, value='Difference from National Mean')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_DIFF_COL, end_column=IteItem.CA3_DIFF_COL)
	ws.cell(row=row, column=IteItem.CBY_MISSED_COL, value='% Missed')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_MISSED_COL, end_column=IteItem.CA3_MISSED_COL)
	ws.cell(row=row, column=IteItem.CBY_DIFFICIENT_COL, value='Difficient Area')
	ws.merge_cells(start_row=row, end_row=row, start_column=IteItem.CBY_DIFFICIENT_COL, end_column=IteItem.CA3_DIFFICIENT_COL)

	for section in sections:
		end_row = section.write_xlsx_rows(ws, row + 1)
		row_ranges.append((row, end_row))
		row = end_row

	data_ranges = get_data_ranges(row_ranges)
	IteSection.write_xlsx_headings(ws, row + 1)
	write_xlsx_summary(ws, data_ranges, row + 2)
	add_conditional_formatting(ws, data_ranges)
	add_styles(ws, data_ranges)

	wb.save(outpath)

def add_styles(worksheet, data_ranges):
	subheading = NamedStyle(name='subheading')
	subheading.font = Font(name='Calibri', size=11)
	subheading.fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
	subheading.number_format = '0%'

	for range_start, _ in data_ranges:
		for row in worksheet['{}{}:{}{}'.format(
			get_column_letter(IteItem.KEYWORD_COL),
			range_start - 2,
			get_column_letter(IteItem.CA3_DIFFICIENT_COL),
			range_start - 1
		)]:
			for cell in row:
				cell.style = subheading

	# summary subheadings
	for row in worksheet['{}{}:{}{}'.format(
		get_column_letter(IteItem.KEYWORD_COL),
		data_ranges[-1][1] + 1,
		get_column_letter(IteItem.CA3_DIFFICIENT_COL),
		data_ranges[-1][1] + 2
	)]:
		for cell in row:
			cell.style = subheading

	# vertical separators
	for col in IteItem.SEPARATOR_COLS + [IteItem.A_OR_B_COL]:
		for row in worksheet.iter_rows(min_row=DATA_START_ROW):
			row[col - 1].style = subheading

	heading = NamedStyle(name='heading')
	heading.font = Font(name='Calibri', size=11, bold=True)
	heading.alignment = Alignment(horizontal='center')

	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_TOTAL_COL).style = heading
	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_DIFF_COL).style = heading
	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_MISSED_COL).style = heading
	worksheet.cell(row=DATA_START_ROW, column=IteItem.CBY_DIFFICIENT_COL).style = heading

	column_ranges = [
		(IteItem.CBY_TOTAL_COL, IteItem.CA3_COL),
		(IteItem.CBY_DIFF_COL, IteItem.CA3_DIFF_COL),
		(IteItem.CBY_MISSED_COL, IteItem.CA3_MISSED_COL)
	]

	data_ranges.append([
		(data_ranges[-1][1] + 3),
		(data_ranges[-1][1] + 5)
	])

	for start_col, end_col in column_ranges:
		for start_row, end_row in data_ranges:
			for row in worksheet['{}{}:{}{}'.format(
				get_column_letter(start_col),
				start_row,
				get_column_letter(end_col),
				end_row
			)]:
				for cell in row:
					cell.style = 'Percent'

	for cell in [
		DIFF_GREAT_CELL,
		DIFF_GOOD_CELL,
		DIFF_BAD_CELL,
		DIFF_VERY_BAD_CELL,
		MISSED_GOOD_CELL,
		MISSED_WARNING_CELL,
		MISSED_BAD_CELL
	]:
		worksheet[cell].style = 'Percent'

def add_conditional_formatting(worksheet, data_ranges):
	original_range = get_ranges(IteItem.CBY_TOTAL_COL, data_ranges,
		end_col=IteItem.CA3_COL, separator=' ')
	diff_range = get_ranges(IteItem.CBY_DIFF_COL, data_ranges,
		end_col=IteItem.CA3_DIFF_COL, separator=' ')
	missed_range = get_ranges(IteItem.CBY_MISSED_COL, data_ranges,
		end_col=IteItem.CA3_MISSED_COL, separator=' ')

	dark_red_fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
	light_red_fill = PatternFill(start_color='ff7d7d', end_color='ff7d7d', fill_type='solid')
	yellow_fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
	green_fill = PatternFill(start_color='55ff55', end_color='55ff55', fill_type='solid')
	blue_fill = PatternFill(start_color='00c2ff', end_color='00c2ff', fill_type='solid')

	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_GREAT, '1'], fill=blue_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_GOOD, DIFF_GREAT], fill=green_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_BAD, DIFF_GOOD], fill=yellow_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=[DIFF_VERY_BAD, DIFF_BAD], fill=light_red_fill))
	worksheet.conditional_formatting.add(diff_range,
		CellIsRule(operator='between', formula=['-1', DIFF_VERY_BAD], fill=dark_red_fill))


	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=['0', MISSED_GOOD], fill=green_fill))
	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=[MISSED_GOOD, MISSED_WARNING], fill=yellow_fill))
	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=[MISSED_WARNING, MISSED_BAD], fill=light_red_fill))
	worksheet.conditional_formatting.add(missed_range,
		CellIsRule(operator='between', formula=[MISSED_BAD, '1'], fill=dark_red_fill))

def write_xlsx_summary(worksheet, data_ranges, row):
	OVERALL_ROW = row
	ADVANCED_ROW = row + 1
	BASIC_ROW = row + 2

	worksheet.cell(row=OVERALL_ROW, column=IteItem.KEYWORD_COL, value='Overall average')
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_TOTAL_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_TOTAL_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_COL, data_ranges)))

	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_DIFF_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_DIFF_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_DIFF_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_DIFF_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_DIFF_COL, data_ranges)))

	worksheet.cell(row=OVERALL_ROW, column=IteItem.CBY_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CBY_MISSED_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA1_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA1_MISSED_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA2_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA2_MISSED_COL, data_ranges)))
	worksheet.cell(row=OVERALL_ROW, column=IteItem.CA3_MISSED_COL,
		value='=AVERAGE({})'.format(get_ranges(IteItem.CA3_MISSED_COL, data_ranges)))

	worksheet.cell(row=ADVANCED_ROW, column=1, value='Advanced question average')
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_TOTAL_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_DIFF_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CBY_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA1_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA2_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=ADVANCED_ROW, column=IteItem.CA3_MISSED_COL,
		value='=AVERAGEIF({},"=A",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=BASIC_ROW, column=1, value='Advanced question average')
	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_TOTAL_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_TOTAL_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_DIFF_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_DIFF_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

	worksheet.cell(row=BASIC_ROW, column=IteItem.CBY_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CBY_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA1_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA1_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA2_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA2_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)
	worksheet.cell(row=BASIC_ROW, column=IteItem.CA3_MISSED_COL,
		value='=AVERAGEIF({},"=B",{})'.format(
			get_range(IteItem.A_OR_B_COL, data_ranges[0][0], data_ranges[-1][-1], absolute_col=True),
			get_range(IteItem.CA3_MISSED_COL, data_ranges[0][0], data_ranges[-1][-1])
		)
	)

def main():
	labels, body = extract('/home/mischka/Downloads/ite and basic stuff/ITE_ProgramItem_156002.txt')

	# dump_csv(labels, body, './output/2017-ite.csv')
	sections = extract_sections(body)
	dump_section_csv(sections, './output/2017-ite-sections.csv')
	dump_section_xlsx(sections, './output/2017-ite-sections.xlsx')


if __name__ == '__main__':
	main()
